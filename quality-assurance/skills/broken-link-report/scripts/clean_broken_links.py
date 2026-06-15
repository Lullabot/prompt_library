#!/usr/bin/env python3
"""
clean_broken_links.py

Filters a Screaming Frog "All Outlinks" CSV down to legitimate broken links and
writes a formatted Excel workbook. The goal is maximum signal, minimum noise:
strip out successes, redirects, links the crawler was simply blocked from
following, and email-obfuscation URLs, leaving only links that are actually broken.

Usage:
    python3 clean_broken_links.py <outlinks.csv> [output.xlsx]

If no output path is given, the workbook is written next to the input file as
    <YYYY-MM-DD>-broken-links.xlsx

The workbook has two sheets:
    Summary  - one row per broken target URL, with how many pages link to it
               (this is where the signal lives; start here)
    Detail   - every broken-link row, so you can trace each break to its source page
"""

import csv
import sys
import os
from datetime import date
from collections import defaultdict
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


def clean_cell(value):
    """Strip control characters Excel refuses to store. Some crawled URLs
    (malformed Teams/meeting links, for instance) carry stray control bytes
    that would otherwise crash the workbook write."""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


# --- Filtering rules ---------------------------------------------------------
# These rules are deliberately host-agnostic so the script works on any site's
# export, not just the one it was first written for.

# Not broken: 200 OK plus every redirect class. A redirect resolves somewhere,
# so it isn't a broken link (audit redirect chains separately if you care).
NON_BROKEN_CODES = {"200", "301", "302", "303", "307", "308"}

# Cloudflare email-obfuscation links. Screaming Frog reports these as 404s, but
# they aren't real pages -- they're spam protection that Cloudflare rewrites at
# render time. The path is the same on every Cloudflare-fronted site, so we match
# the path fragment rather than a specific host.
EMAIL_PROTECTION_FRAGMENT = "/cdn-cgi/l/email-protection"

# Columns pulled from the export for the Detail sheet. The rest of Screaming
# Frog's columns (Size, Transferred, Rel, Alt Text, etc.) are noise for triage.
KEEP_COLS = ["From", "To", "Anchor Text", "Status Code", "Status", "Type", "Link Position"]

# Advisory flag text. A 403 usually means the destination server refused the
# crawler (anti-bot protection) rather than the resource being genuinely gone --
# so these stay in the broken list but get tagged for a human to eyeball.
BOT_BLOCK_FLAG = "Likely bot-blocked - verify manually"


def flag_for(code):
    return BOT_BLOCK_FLAG if code == "403" else ""


def is_noise(row):
    """True if the row is NOT a legitimate broken link and should be dropped."""
    code = (row.get("Status Code") or "").strip()
    status = (row.get("Status") or "").strip().lower()
    to = (row.get("To") or "").strip()

    if EMAIL_PROTECTION_FRAGMENT in to:
        return True
    if code in NON_BROKEN_CODES:
        return True
    # Status code 0 means the crawler never got an HTTP response. "Blocked by
    # robots.txt" is the crawler being told to stay out -- uncrawlable, not broken.
    # Other code-0 reasons (DNS lookup failed, connection refused/timeout) ARE
    # genuine breaks, so only robots.txt is filtered here.
    if code == "0" and "robots" in status:
        return True
    return False


def load_broken_rows(infile):
    rows = []
    with open(infile, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames or "Status Code" not in reader.fieldnames:
            raise ValueError(
                "This doesn't look like a Screaming Frog outlinks export "
                "(no 'Status Code' column). Expected columns include "
                "Type, From, To, Anchor Text, Status Code, Status, ..."
            )
        for row in reader:
            if is_noise(row):
                continue
            out = {c: (row.get(c) or "") for c in KEEP_COLS}
            out["Flag"] = flag_for(out["Status Code"].strip())
            rows.append(out)
    # Real server errors first (by status code), then by target URL.
    rows.sort(key=lambda r: (r["Status Code"], r["To"]))
    return rows


def build_summary(rows):
    """Collapse to one row per broken target URL. This is what makes a giant
    export legible: a link broken in a site-wide footer shows up once with a
    high page count, instead of hundreds of near-identical rows."""
    agg = defaultdict(lambda: {"linked": 0, "sources": set()})
    meta = {}
    for r in rows:
        to = r["To"]
        agg[to]["linked"] += 1
        agg[to]["sources"].add(r["From"])
        meta[to] = (r["Status Code"], r["Status"])
    summary = []
    for to, data in agg.items():
        code, status = meta[to]
        summary.append({
            "To": to,
            "Host": urlparse(to).netloc,
            "Status Code": code,
            "Status": status,
            "Times Linked": data["linked"],
            "Source Pages": len(data["sources"]),
            "Flag": flag_for(code.strip()),
        })
    # Most-linked breaks first -- they're the highest-impact fixes.
    summary.sort(key=lambda r: (-r["Times Linked"], r["Status Code"]))
    return summary


# --- Excel output ------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
# Soft amber so flagged rows are visible without screaming "error".
FLAG_FILL = PatternFill("solid", fgColor="FFF2CC")


def write_sheet(ws, fieldnames, rows):
    ws.append(fieldnames)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    flag_idx = fieldnames.index("Flag") if "Flag" in fieldnames else None
    for r in rows:
        ws.append([clean_cell(r.get(c, "")) for c in fieldnames])
        # Tint the whole row of any flagged (403) entry for quick scanning.
        if flag_idx is not None and r.get("Flag"):
            for cell in ws[ws.max_row]:
                cell.fill = FLAG_FILL
    ws.freeze_panes = "A2"
    # Auto-ish column widths, capped so long URLs don't blow the layout out.
    for i, name in enumerate(fieldnames, start=1):
        longest = max([len(str(name))] + [len(str(r.get(name, ""))) for r in rows] or [0])
        ws.column_dimensions[get_column_letter(i)].width = min(max(longest + 2, 10), 80)
    if rows:
        ws.auto_filter.ref = ws.dimensions


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    infile = sys.argv[1]
    if len(sys.argv) >= 3:
        outfile = sys.argv[2]
    else:
        outdir = os.path.dirname(os.path.abspath(infile))
        outfile = os.path.join(outdir, f"{date.today().isoformat()}-broken-links.xlsx")

    rows = load_broken_rows(infile)
    summary = build_summary(rows)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    write_sheet(ws_summary,
                ["To", "Host", "Status Code", "Status", "Times Linked", "Source Pages", "Flag"],
                summary)
    ws_detail = wb.create_sheet("Detail")
    write_sheet(ws_detail, KEEP_COLS + ["Flag"], rows)
    wb.save(outfile)

    flagged = sum(1 for r in summary if r["Flag"])
    print(f"Broken-link rows: {len(rows)}  |  unique broken URLs: {len(summary)}  |  flagged (403): {flagged}")
    print(f"Workbook written: {outfile}")


if __name__ == "__main__":
    main()
