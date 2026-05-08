"""Output formatting — produces audit-template-formatted CSV or Excel files."""

import csv
import io
from pathlib import Path

import pandas as pd


# Pages audit template column structure
PAGES_CATEGORY_ROW = [
    "Page information (pulled and provided)", "", "", "", "", "",
    "Review criteria", "", "", "", "",
    "Decisions & comments", "", "",
    "Responsbilities", "",
    "For Lullabot",
]

PAGES_HEADER_ROW = [
    "Page title\nThe title of this page.",
    "Page URL\nThe URL to this page.",
    "Redirect, Plaintext\nConvert this to the main page column for agency once cleanup is done",
    "GA page views\nGoogle Analytics data \n(1 year prior to scan).",
    "Reading level: Score",
    "Reading level: Grade\nAim for Normal or easier (Normal equals 8th Grade).",
    "Audience\nWho is this content for and how will it serve them? Consider the key audience goals and tasks.",
    "Up to date\nContent is current and up-to-date, meeting current needs and requirements.",
    "Accuracy\nContent should be accurate, free of errors, and leads audiences to the correct next action.",
    "Uniqueness\nContent should be unique - a single page that doesn't exist somewhere else in the same or similar format.",
    "High quality\nGrade the page on a scale of 1-5 (1 = low, 5 = high), e.g. does it meet your business and audience goals? Does it align with brand messaging?",
    "Keep, edit, delete or consolidate? \nIf consolidating, note what page to combine it with.",
    "Required by law\nThis content is required to be on the website",
    "Notes, questions, comments\nLeave any notes, gut reactions, thoughts about the content, its value, or quality.",
    "Reviewed by",
    "Stakeholder (if applicable)",
    "Notes",
]

# Files audit template column structure
FILES_CATEGORY_ROW = [
    "File information (pulled and provided)", "", "", "", "", "", "", "",
    "Review criteria", "", "", "", "",
    "Decisions & comments", "", "",
    "Responsbilities", "",
    "For Lullabot",
]

FILES_HEADER_ROW = [
    "File title\nThe title of this file",
    "File URL\nThe URL to this file.",
    "File URL without formatting \n(http(s), www. & trailing slashes)",
    "File type\nThe type of file",
    "Anchor text\nThe text of the link to this file.",
    "Source\nThe page that has or links to this file.",
    "Alt text\nThe image alt text",
    "Size\nSize of this file (bytes).",
    "Audience\nWho is this content for and how will it serve them? Consider the key audience goals and tasks.",
    "Up to date\nContent is current and up-to-date, meeting current needs and requirements.",
    "Accuracy\nContent should be accurate, free of errors, and leads audiences to the correct next action.",
    "Uniqueness\nContent should be unique - a single page that doesn't exist somewhere else in the same or similar format.",
    "High quality\nGrade the page on a scale of 1-5 (1 = low, 5 = high), e.g. does it meet your business and audience goals? Does it align with brand messaging?",
    "Keep, edit, delete or consolidate? \nIf consolidating, note what page to combine it with.",
    "Required by law\nThis content is required to be on the website",
    "Notes, questions, comments\nLeave any notes, gut reactions, thoughts about the content, its value, or quality.",
    "Reviewed by",
    "Stakeholder (if applicable)",
    "Notes",
]


def _format_value(val):
    """Format a value for output, converting NaN/None to empty string."""
    if pd.isna(val):
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val)


def format_pages_csv(df: pd.DataFrame, output_path: str):
    """Write pages DataFrame to a CSV matching the audit template format."""
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(PAGES_CATEGORY_ROW)
        writer.writerow(PAGES_HEADER_ROW)

        for _, row in df.iterrows():
            # Build the redirect full URL from normalized URL
            redirect_url = _format_value(row.get("redirect_url", ""))
            if not redirect_url:
                # Reconstruct full URL for the redirect column
                norm = str(row["normalized_url"]).split("\n")[0]
                redirect_url = f"https://{norm}" if norm else ""

            data_row = [
                _format_value(row.get("page_title", "")),
                _format_value(row.get("normalized_url", "")),
                redirect_url,
                _format_value(row.get("ga_views", "")),
                _format_value(row.get("reading_score", "")),
                _format_value(row.get("reading_grade", "")),
                "",  # Audience
                "",  # Up to date
                "",  # Accuracy
                "",  # Uniqueness
                "",  # High quality
                "",  # Keep/edit/delete
                "FALSE",  # Required by law
                "",  # Notes
                "",  # Reviewed by
                "",  # Stakeholder
                "",  # Notes (Lullabot)
            ]
            writer.writerow(data_row)

    print(f"Wrote pages audit sheet: {output_path}")


def format_files_csv(df: pd.DataFrame, output_path: str):
    """Write files DataFrame to a CSV matching the audit template format."""
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(FILES_CATEGORY_ROW)
        writer.writerow(FILES_HEADER_ROW)

        for _, row in df.iterrows():
            data_row = [
                _format_value(row.get("file_title", "")),
                _format_value(row.get("file_url", "")),
                _format_value(row.get("normalized_url", "")),
                _format_value(row.get("file_type", "")),
                _format_value(row.get("anchor_text", "")),
                _format_value(row.get("source_urls", "")),
                _format_value(row.get("alt_text", "")),
                _format_value(row.get("size", "")),
                "",  # Audience
                "",  # Up to date
                "",  # Accuracy
                "",  # Uniqueness
                "",  # High quality
                "",  # Keep/edit/delete
                "FALSE",  # Required by law
                "",  # Notes
                "",  # Reviewed by
                "",  # Stakeholder
                "",  # Notes (Lullabot)
            ]
            writer.writerow(data_row)

    print(f"Wrote files audit sheet: {output_path}")


def format_pages_excel(df: pd.DataFrame, output_path: str):
    """Write pages DataFrame to an Excel file with formatted headers."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory All Pages"

    # Write category header
    for col, val in enumerate(PAGES_CATEGORY_ROW, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.font = Font(bold=True)

    # Write column headers
    for col, val in enumerate(PAGES_HEADER_ROW, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(bold=True)

    # Write data
    for row_idx, (_, row) in enumerate(df.iterrows(), 3):
        redirect_url = _format_value(row.get("redirect_url", ""))
        if not redirect_url:
            norm = str(row["normalized_url"]).split("\n")[0]
            redirect_url = f"https://{norm}" if norm else ""

        values = [
            _format_value(row.get("page_title", "")),
            _format_value(row.get("normalized_url", "")),
            redirect_url,
            _format_value(row.get("ga_views", "")),
            _format_value(row.get("reading_score", "")),
            _format_value(row.get("reading_grade", "")),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if "\n" in str(val):
                cell.alignment = Alignment(wrap_text=True)

        # Required by law default
        ws.cell(row=row_idx, column=13, value="FALSE")

    wb.save(output_path)
    print(f"Wrote pages audit sheet (Excel): {output_path}")


def format_files_excel(df: pd.DataFrame, output_path: str):
    """Write files DataFrame to an Excel file with formatted headers."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "All Files"

    for col, val in enumerate(FILES_CATEGORY_ROW, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.font = Font(bold=True)

    for col, val in enumerate(FILES_HEADER_ROW, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(bold=True)

    for row_idx, (_, row) in enumerate(df.iterrows(), 3):
        values = [
            _format_value(row.get("file_title", "")),
            _format_value(row.get("file_url", "")),
            _format_value(row.get("normalized_url", "")),
            _format_value(row.get("file_type", "")),
            _format_value(row.get("anchor_text", "")),
            _format_value(row.get("source_urls", "")),
            _format_value(row.get("alt_text", "")),
            _format_value(row.get("size", "")),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if "\n" in str(val):
                cell.alignment = Alignment(wrap_text=True)

        ws.cell(row=row_idx, column=15, value="FALSE")

    wb.save(output_path)
    print(f"Wrote files audit sheet (Excel): {output_path}")
