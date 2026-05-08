#!/usr/bin/env python3
"""Generate az-gov-data-dictionary.xlsx from Drupal 9 config YAML files in sync/."""

import argparse
import os
import glob
import re
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

SYNC_DIR = 'sync'
OUTPUT_CT = 'az-gov-content-types.xlsx'
OUTPUT_TV = 'az-gov-taxonomy-vocabularies.xlsx'
OUTPUT_BT = 'az-gov-block-types.xlsx'
OUTPUT_MT = 'az-gov-media-types.xlsx'
OUTPUT_PT = 'az-gov-paragraph-types.xlsx'

# Colors matched from Template_ Data Dictionary.xlsx
CLR_BLUE  = '4A86E8'  # Row 1 label headers
CLR_GREY  = 'B7B7B7'  # Row 2 entity name
CLR_LGREY = 'D9D9D9'  # Row 3 column headers
CLR_LIST  = 'EFEFEF'  # List sheet headers

FIELD_TYPE_MAP = {
    'string':            'Text, short',
    'string_long':       'Text, long',
    'text_with_summary': 'Text, formatted, long',
    'text_long':         'Text, formatted, long',
    'text':              'Text, formatted',
    'entity_reference':  'Entity reference',
    'daterange':         'Date range',
    'datetime':          'Date/time',
    'boolean':           'Boolean',
    'link':              'Link',
    'telephone':         'Telephone',
    'integer':           'Number, integer',
    'float':             'Number, decimal',
    'decimal':           'Number, decimal',
    'list_string':       'List, text',
    'list_integer':      'List, integer',
    'list_float':        'List, decimal',
    'file':              'File',
    'image':             'Image',
    'address':           'Address',
    'metatag':           'Meta tags',
    'email':             'Email',
    'uri':               'URI',
    'map':               'Key-value pairs',
}

WIDGET_TYPE_MAP = {
    'string_textfield':                  'Text input',
    'string_textarea':                   'Text area',
    'text_textarea':                     'Text area',
    'text_textarea_with_summary':        'Text area with summary',
    'entity_reference_autocomplete':     'Autocomplete',
    'entity_reference_autocomplete_tags':'Autocomplete (tags)',
    'media_library_widget':              'Media library',
    'options_select':                    'Select list',
    'link_default':                      'Link input',
    'boolean_checkbox':                  'Checkbox',
    'telephone_default':                 'Phone input',
    'datetime_default':                  'Date/time input',
    'daterange_default':                 'Date range input',
    'file_generic':                      'File upload',
    'image_image':                       'Image upload',
    'address_default':                   'Address input',
    'metatag_firehose':                  'Meta tags widget',
    'layout_builder_widget':             'Layout builder',
    'email_default':                     'Email input',
    'number':                            'Number input',
    'path':                              'URL alias',
}

# System/admin fields to exclude from the dictionary
SYSTEM_FIELDS = {
    'uid', 'created', 'changed', 'promote', 'sticky', 'status',
    'moderation_state', 'path', 'url_redirects', 'revision_log',
    'langcode', 'content_translation_source', 'content_translation_outdated',
    'revision_uid', 'revision_timestamp', 'revision_log_message',
}

DETAIL_COLUMNS = [
    'Property', 'Machine name', 'Data type', 'Values',
    'Cardinality', 'Required', 'Rules',
    'Form display:\nField widget type', 'Form display:\nCharacter limit',
    'Form display: \nHelp text', 'Notes', 'D7 Field',
]

# Column widths matched from template
DETAIL_COL_WIDTHS = [22, 30, 22.63, 20, 14, 7.88, 12.38, 15.25, 12.13, 58.63, 40, 15]


def load_yaml(path):
    with open(path) as f:
        return yaml.safe_load(f)


def strip_html(text):
    """Remove HTML tags and decode basic entities from a string."""
    if not text:
        return text
    text = re.sub(r'<[^>]+>', '', text)
    text = text.replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>').replace('&nbsp;', ' ').replace('&#039;', "'").replace('&quot;', '"')
    return text.strip()


def solid_fill(color):
    return PatternFill('solid', fgColor=color)


def get_label_for_bundle(entity_type, bundle):
    """Return the human-readable name for a bundle."""
    paths = {
        'node':          f'node.type.{bundle}.yml',
        'taxonomy_term': f'taxonomy.vocabulary.{bundle}.yml',
        'block_content': f'block_content.type.{bundle}.yml',
        'media':         f'media.type.{bundle}.yml',
        'paragraph':     f'paragraphs.paragraphs_type.{bundle}.yml',
    }
    path = os.path.join(SYNC_DIR, paths.get(entity_type, ''))
    if os.path.exists(path):
        d = load_yaml(path)
        return d.get('name') or d.get('label') or d.get('vid') or bundle
    return bundle


def get_field_storage(entity_type, field_name):
    path = os.path.join(SYNC_DIR, f'field.storage.{entity_type}.{field_name}.yml')
    if os.path.exists(path):
        return load_yaml(path)
    return None


def get_form_display_data(entity_type, bundle):
    path = os.path.join(SYNC_DIR, f'core.entity_form_display.{entity_type}.{bundle}.default.yml')
    if os.path.exists(path):
        return load_yaml(path)
    return None


def format_cardinality(cardinality):
    if cardinality == 1:
        return 'Single'
    if cardinality == -1:
        return 'Multiple'
    return f'Up to {cardinality}'


def format_values(field_data, storage_data):
    """Build a human-readable description of allowed values for a field."""
    field_type = field_data.get('field_type', '')
    settings = field_data.get('settings', {})
    handler = settings.get('handler', '')
    handler_settings = settings.get('handler_settings', {})
    target_bundles = handler_settings.get('target_bundles') or {}

    if field_type == 'entity_reference' and target_bundles:
        if 'taxonomy_term' in handler:
            names = [get_label_for_bundle('taxonomy_term', vid) for vid in target_bundles]
        elif 'media' in handler:
            names = [get_label_for_bundle('media', mt) for mt in target_bundles]
        elif 'node' in handler:
            names = [get_label_for_bundle('node', b) for b in target_bundles]
        elif 'block_content' in handler:
            names = [get_label_for_bundle('block_content', b) for b in target_bundles]
        elif 'paragraph' in handler:
            names = [get_label_for_bundle('paragraph', b) for b in target_bundles]
        else:
            names = list(target_bundles.keys())
        return ', '.join(names)

    if field_type in ('list_string', 'list_integer', 'list_float') and storage_data:
        allowed = storage_data.get('settings', {}).get('allowed_values', [])
        labels = []
        for v in allowed:
            if isinstance(v, dict):
                labels.append(v.get('label') or v.get('value', ''))
            else:
                labels.append(str(v))
        if labels:
            return '\n'.join(f'• {l}' for l in labels)

    if field_type == 'boolean':
        return 'On / Off'

    return ''


def format_widget(widget_type, cardinality):
    if widget_type == 'options_buttons':
        return 'Radio buttons' if cardinality == 1 else 'Checkboxes'
    return WIDGET_TYPE_MAP.get(widget_type, widget_type or '')


def get_custom_fields_sorted(entity_type, bundle, form_display_data):
    """Return [(field_name, field_data)] for all non-system custom fields, ordered by form weight."""
    pattern = os.path.join(SYNC_DIR, f'field.field.{entity_type}.{bundle}.*.yml')
    fields = {}
    for path in glob.glob(pattern):
        d = load_yaml(path)
        fname = d.get('field_name', '')
        if fname and fname not in SYSTEM_FIELDS:
            fields[fname] = d

    fd_content = (form_display_data or {}).get('content', {})

    def weight(item):
        return fd_content.get(item[0], {}).get('weight', 9999)

    return sorted(fields.items(), key=weight)


def write_entity_sheet(wb, sheet_name, entity_label, entity_type_label, entity_desc, fields, form_display_data):
    ws = wb.create_sheet(title=sheet_name[:31])
    fd_content = (form_display_data or {}).get('content', {})
    ncols = len(DETAIL_COLUMNS)

    # Row 1 — blue header labels: Name | Type | Status
    for col, h in enumerate(['Name', 'Type', 'Status'], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = solid_fill(CLR_BLUE)
        c.alignment = Alignment(vertical='top')

    # Row 2 — entity metadata row (grey)
    ws.cell(row=2, column=1, value=entity_label)
    ws.cell(row=2, column=2, value=entity_type_label)
    ws.cell(row=2, column=3, value='')
    for col in range(1, ncols + 1):
        ws.cell(row=2, column=col).fill = solid_fill(CLR_GREY)
    ws.cell(row=2, column=1).font = Font(bold=True, size=12)

    # Row 3 — column headers (light grey)
    for col, h in enumerate(DETAIL_COLUMNS, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = solid_fill(CLR_LGREY)
        c.alignment = Alignment(wrap_text=True, vertical='top')

    # Data rows
    data_row = 4
    for field_name, field_data in fields:
        entity_type = field_data.get('entity_type', 'node')
        storage = get_field_storage(entity_type, field_name)
        raw_type = field_data.get('field_type', '')
        cardinality = (storage or {}).get('cardinality', 1)

        widget_info = fd_content.get(field_name, {})
        widget_type = widget_info.get('type', '')
        if storage:
            char_limit = storage.get('settings', {}).get('max_length') or ''
        elif raw_type == 'string':
            char_limit = 255  # Drupal default max for string fields
        else:
            char_limit = ''

        row_vals = [
            field_data.get('label', ''),
            field_name,
            FIELD_TYPE_MAP.get(raw_type, raw_type),
            format_values(field_data, storage),
            format_cardinality(cardinality),
            'x' if field_data.get('required') else '',
            '',  # Rules — manual
            format_widget(widget_type, cardinality),
            char_limit if char_limit else '',
            strip_html(field_data.get('description', '')),
            '',  # Notes — manual
            '',  # D7 Field — manual
        ]
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=data_row, column=col, value=val)
            c.alignment = Alignment(wrap_text=True, vertical='top')
        data_row += 1

    # Notes block
    notes_row = data_row + 1
    ws.cell(row=notes_row, column=1, value='Notes').font = Font(bold=True)
    if entity_desc:
        ws.cell(row=notes_row + 1, column=1, value=entity_desc).alignment = Alignment(wrap_text=True, vertical='top')

    # Column widths and row heights
    for col, width in enumerate(DETAIL_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 16.5
    ws.row_dimensions[2].height = 23.25
    ws.row_dimensions[3].height = 30

    return ws


def write_list_sheet(wb, title, headers, rows, col_widths, link_col=None, link_names=None):
    ws = wb.create_sheet(title=title)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, size=12)
        c.fill = solid_fill(CLR_LIST)
        c.alignment = Alignment(vertical='top')
    for row_idx, row_data in enumerate(rows, 2):
        for col, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=val).alignment = Alignment(wrap_text=True, vertical='top')
        if link_col and link_names:
            target_sheet = link_names[row_idx - 2]
            col_letter = get_column_letter(link_col)
            c = ws.cell(row=row_idx, column=link_col)
            c.hyperlink = Hyperlink(
                ref=f'{col_letter}{row_idx}',
                location=f"'{target_sheet}'!A1",
            )
            c.font = Font(color='1155CC', underline='single')
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 20
    return ws


def safe_name(prefix, label):
    # Excel sheet names: 31-char max, no : \ / ? * [ ]
    s = f'{prefix} {label}'
    for ch in r'\/?*[]':
        s = s.replace(ch, '')
    return s[:31]


def make_workbook():
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def main():
    parser = argparse.ArgumentParser(description='Generate Drupal data dictionary Excel files.')
    parser.add_argument('--sync-dir', default='sync',
                        help='Path to Drupal config sync directory, relative to CWD (default: sync)')
    parser.add_argument('--prefix', default='az-gov',
                        help='Prefix for output filenames (default: az-gov)')
    args = parser.parse_args()

    global SYNC_DIR
    SYNC_DIR = args.sync_dir
    p = args.prefix
    output_ct = f'{p}-content-types.xlsx'
    output_tv = f'{p}-taxonomy-vocabularies.xlsx'
    output_bt = f'{p}-block-types.xlsx'
    output_mt = f'{p}-media-types.xlsx'
    output_pt = f'{p}-paragraph-types.xlsx'

    # ── Content Types ──────────────────────────────────────────────────────────
    wb_ct = make_workbook()
    ct_files = glob.glob(os.path.join(SYNC_DIR, 'node.type.*.yml'))
    content_types = sorted([load_yaml(p) for p in ct_files], key=lambda x: x.get('name', ''))
    ct_list_rows = []
    ct_sheet_names = []

    for ct in content_types:
        name   = ct.get('name', '')
        bundle = ct.get('type', '')
        desc   = ct.get('description', '')

        fd_data = get_form_display_data('node', bundle)
        fd_content = (fd_data or {}).get('content', {})

        override_path = os.path.join(SYNC_DIR, f'core.base_field_override.node.{bundle}.title.yml')
        title_label = 'Title'
        title_desc = ''
        if os.path.exists(override_path):
            ov = load_yaml(override_path)
            title_label = ov.get('label', 'Title')
            title_desc  = ov.get('description', '')

        title_entry = ('title', {
            'label': title_label,
            'field_name': 'title',
            'field_type': 'string',
            'required': True,
            'description': title_desc,
            'entity_type': 'node',
        })

        custom_fields = get_custom_fields_sorted('node', bundle, fd_data)

        title_weight = fd_content.get('title', {}).get('weight', -9999)
        all_fields = []
        title_inserted = False
        for fname, fdata in custom_fields:
            w = fd_content.get(fname, {}).get('weight', 9999)
            if not title_inserted and title_weight <= w:
                all_fields.append(title_entry)
                title_inserted = True
            all_fields.append((fname, fdata))
        if not title_inserted:
            all_fields.append(title_entry)

        clean_desc = strip_html(desc)
        sheet_name = safe_name('CT', name)
        write_entity_sheet(wb_ct, sheet_name, name, 'Content type', clean_desc, all_fields, fd_data)
        ct_list_rows.append([name, clean_desc, '', '', bundle, ''])
        ct_sheet_names.append(sheet_name)

    write_list_sheet(
        wb_ct, 'Content Type List',
        ['Content Type', 'Description', 'Examples', 'Notes', 'Machine Name', 'URL Pattern'],
        ct_list_rows,
        [25, 60, 20, 20, 20, 20],
        link_col=5, link_names=ct_sheet_names,
    )
    wb_ct.move_sheet('Content Type List', offset=-wb_ct.sheetnames.index('Content Type List'))
    wb_ct.save(output_ct)
    print(f'Saved: {output_ct} ({len(wb_ct.sheetnames)} sheets)')

    # ── Taxonomy Vocabularies ──────────────────────────────────────────────────
    wb_tv = make_workbook()
    tv_files = glob.glob(os.path.join(SYNC_DIR, 'taxonomy.vocabulary.*.yml'))
    vocabs = sorted([load_yaml(p) for p in tv_files], key=lambda x: x.get('name', ''))
    tv_list_rows = []
    tv_sheet_names = []

    for vocab in vocabs:
        name = vocab.get('name', '')
        vid  = vocab.get('vid', '')
        desc = vocab.get('description', '')
        sheet_name = safe_name('TV', name)
        tv_sheet_names.append(sheet_name)
        tv_list_rows.append([name, vid, strip_html(desc), ''])

        custom_field_files = glob.glob(os.path.join(SYNC_DIR, f'field.field.taxonomy_term.{vid}.*.yml'))
        fd_data = get_form_display_data('taxonomy_term', vid)

        base_fields = [
            ('name', {
                'label': 'Name', 'field_name': 'name', 'field_type': 'string',
                'required': True, 'description': '', 'entity_type': 'taxonomy_term',
            }),
            ('description', {
                'label': 'Description', 'field_name': 'description', 'field_type': 'text_long',
                'required': False, 'description': '', 'entity_type': 'taxonomy_term',
            }),
        ]
        custom_fields = []
        for path in custom_field_files:
            d = load_yaml(path)
            fname = d.get('field_name', '')
            if fname and fname not in SYSTEM_FIELDS:
                custom_fields.append((fname, d))

        write_entity_sheet(
            wb_tv, sheet_name, name, 'Taxonomy vocabulary', strip_html(desc),
            base_fields + custom_fields, fd_data,
        )

    write_list_sheet(
        wb_tv, 'Taxonomy Vocabulary List',
        ['Vocabulary', 'Machine Name', 'Description', 'Notes'],
        tv_list_rows,
        [25, 25, 60, 25],
        link_col=2, link_names=tv_sheet_names,
    )
    wb_tv.move_sheet('Taxonomy Vocabulary List', offset=-wb_tv.sheetnames.index('Taxonomy Vocabulary List'))
    wb_tv.save(output_tv)
    print(f'Saved: {output_tv} ({len(wb_tv.sheetnames)} sheets)')

    # ── Block Content Types ────────────────────────────────────────────────────
    wb_bt = make_workbook()
    bt_files = glob.glob(os.path.join(SYNC_DIR, 'block_content.type.*.yml'))
    block_types = sorted([load_yaml(p) for p in bt_files], key=lambda x: x.get('label', ''))
    bt_list_rows = []
    bt_sheet_names = []

    for bt in block_types:
        label    = bt.get('label', '')
        block_id = bt.get('id', '')
        desc     = bt.get('description', '')

        fd_data = get_form_display_data('block_content', block_id)
        custom_fields = get_custom_fields_sorted('block_content', block_id, fd_data)

        info_entry = ('info', {
            'label': 'Block description',
            'field_name': 'info',
            'field_type': 'string',
            'required': True,
            'description': 'Administrative label for this block (not displayed to site visitors).',
            'entity_type': 'block_content',
        })

        clean_desc = strip_html(desc)
        sheet_name = safe_name('BT', label)
        write_entity_sheet(wb_bt, sheet_name, label, 'Block type', clean_desc, [info_entry] + custom_fields, fd_data)
        bt_list_rows.append([label, clean_desc, block_id, ''])
        bt_sheet_names.append(sheet_name)

    write_list_sheet(
        wb_bt, 'Block Type List',
        ['Block Type', 'Description', 'Machine Name', 'Notes'],
        bt_list_rows,
        [25, 60, 20, 25],
        link_col=3, link_names=bt_sheet_names,
    )
    wb_bt.move_sheet('Block Type List', offset=-wb_bt.sheetnames.index('Block Type List'))
    wb_bt.save(output_bt)
    print(f'Saved: {output_bt} ({len(wb_bt.sheetnames)} sheets)')

    # ── Media Types ────────────────────────────────────────────────────────────
    wb_mt = make_workbook()
    mt_files = glob.glob(os.path.join(SYNC_DIR, 'media.type.*.yml'))
    media_types = sorted([load_yaml(p) for p in mt_files], key=lambda x: x.get('label', ''))
    mt_list_rows = []
    mt_sheet_names = []

    for mt in media_types:
        label    = mt.get('label', '')
        mid      = mt.get('id', '')
        source   = mt.get('source', '')
        raw_desc = mt.get('description', '')
        clean_desc = strip_html(raw_desc)

        fd_data = get_form_display_data('media', mid)
        custom_fields = get_custom_fields_sorted('media', mid, fd_data)

        # Media source field is the "primary" field (e.g. field_media_image for images)
        source_field_name = mt.get('source_configuration', {}).get('source_field', '')
        name_entry = ('name', {
            'label': 'Name', 'field_name': 'name', 'field_type': 'string',
            'required': True, 'description': 'Administrative name for this media item.',
            'entity_type': 'media',
        })

        sheet_name = safe_name('MT', label)
        # Prepend source to description so it's visible in the Notes block
        full_desc = f'Source plugin: {source}\n{clean_desc}' if source else clean_desc
        write_entity_sheet(wb_mt, sheet_name, label, 'Media type', full_desc, [name_entry] + custom_fields, fd_data)
        mt_list_rows.append([label, clean_desc, mid, source, ''])
        mt_sheet_names.append(sheet_name)

    write_list_sheet(
        wb_mt, 'Media Type List',
        ['Media Type', 'Description', 'Machine Name', 'Source', 'Notes'],
        mt_list_rows,
        [25, 55, 20, 20, 25],
        link_col=3, link_names=mt_sheet_names,
    )
    wb_mt.move_sheet('Media Type List', offset=-wb_mt.sheetnames.index('Media Type List'))
    wb_mt.save(output_mt)
    print(f'Saved: {output_mt} ({len(wb_mt.sheetnames)} sheets)')

    # ── Paragraph Types ────────────────────────────────────────────────────────
    wb_pt = make_workbook()
    pt_files = glob.glob(os.path.join(SYNC_DIR, 'paragraphs.paragraphs_type.*.yml'))
    para_types = sorted([load_yaml(p) for p in pt_files], key=lambda x: x.get('label', ''))
    pt_list_rows = []
    pt_sheet_names = []

    for pt in para_types:
        label  = pt.get('label', '')
        pid    = pt.get('id', '')
        desc   = pt.get('description', '')
        clean_desc = strip_html(desc)

        fd_data = get_form_display_data('paragraph', pid)
        custom_fields = get_custom_fields_sorted('paragraph', pid, fd_data)

        sheet_name = safe_name('PT', label)
        write_entity_sheet(wb_pt, sheet_name, label, 'Paragraph type', clean_desc, custom_fields, fd_data)
        pt_list_rows.append([label, clean_desc, pid, ''])
        pt_sheet_names.append(sheet_name)

    write_list_sheet(
        wb_pt, 'Paragraph Type List',
        ['Paragraph Type', 'Description', 'Machine Name', 'Notes'],
        pt_list_rows,
        [25, 60, 20, 25],
        link_col=3, link_names=pt_sheet_names,
    )
    wb_pt.move_sheet('Paragraph Type List', offset=-wb_pt.sheetnames.index('Paragraph Type List'))
    wb_pt.save(output_pt)
    print(f'Saved: {output_pt} ({len(wb_pt.sheetnames)} sheets)')


if __name__ == '__main__':
    main()
